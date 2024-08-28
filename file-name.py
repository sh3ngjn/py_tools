import os
import openpyxl

# 指定文件夹路径
folder_path = ''  # 请将路径替换为你的实际文件夹路径

# 创建一个新的工作簿和工作表
workbook = openpyxl.Workbook()
sheet = workbook.active

# 设置表头
sheet.cell(row=1, column=1, value='文件名')

# 遍历文件夹中的所有文件
row = 2
for filename in os.listdir(folder_path):
    # 将文件名写入表格
    sheet.cell(row=row, column=1, value=filename)
    row += 1

# 保存Excel文件
excel_path = ''  # 请将路径替换为你要保存Excel文件的路径
workbook.save(excel_path)